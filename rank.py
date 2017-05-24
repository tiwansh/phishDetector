import openpyxl
import time
import titleSearch
import re
import requests
from tld import get_tld
import socket



'''
limit = raw_input("Enter limit : ")
limit = int(limit)
print type(limit)
'''

#Filename.xlsx
book = openpyxl.load_workbook('test1.xlsx')

sheet = book.active
i = 2
s2 = sheet.cell(row = i, column = 1)
'''print s2.value

i = i + 1
s2 = sheet.cell(row = i, column = 1)
print s2.value
'''

phishCount = 0
legitCount = 0
totalTime = 0
errorCount = 0

while 1:
	#print "Here"
	s2 = sheet.cell(row = i, column = 1)
	if s2.value is None:
		break
	url = str(s2.value)


	url = "http://www." + url
	print (url)
	#i = i + 1
	
	'''title = titleSearch.returnTitle(url)
	if(title is not ""):
		print "Title is : " + str(title)
		#flag_titleFound = True
	else:
		print ("Title not found!")
	'''

	start_time = time.time()
	#verdict = srcIPMatch.ipMatches(url, 3)
	try:
		#print "gone"
		#rank nikalo
		#total hyperlinks
		#no. of hyperlinks of same domain
		dictionary = {}
		dictionarylist = []

		dictionary.clear()
		del dictionarylist[:]
		print "Inside Second Module"
		flagLegitimate = False
		try:
			urlIP = socket.gethostbyname(get_tld(url))
		except:
			print "Unable host urlIP"
			#return False
		counter = 0
		r = requests.get(url)

		urls = re.findall('http[s]?://(?:[a-zA-Z]|[0-9]|[$-_@.&+]|[!*\(\),]|(?:%[0-9a-fA-F][0-9a-fA-F]))+', r.content)
		totalLinks = 0
		for link in urls:
			totalLinks = totalLinks + 1
			#for arbitrary links with only words ?
			if "." not in link:
				continue
			pattern = re.compile("\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}")
			if pattern.match(link):
				continue

			#Testing whether the link is an IP ?
			#Trying in case of IP address hosted link example -- http://128.199.202.96/thehindu/public/img/article/theHinduCenter.png
			#print link
			try:
				newurl = get_tld(link)
				#print newurl
				if not newurl in dictionary:
					dictionary[newurl] = 1
				else:
					dictionary[newurl] += 1
			except:
				continue

		print totalLinks
		#Created a new sorted list from dictionary
		dictionarylist = sorted(dictionary, key=dictionary.get, reverse = True)
		matchFreq = 0;
		rankTld = 1;
		for urls in dictionarylist:
			'''if(counter == limit):
				break
			else:
				counter = counter + 1'''
			try:
				urlsIP = socket.gethostbyname(urls)
				print urls + " " + socket.gethostbyname(urls) + " " + urlIP
				if urlIP == urlsIP:
					print "---------------IP's Matched-------------"
					matchFreq = dictionary[urls]
					flagLegitimate = True
					break
			except:
				print urls + " " + "Exception !"
			rankTld = rankTld + 1

	except:
		i = i + 1
		continue

	time_taken = time.time() - start_time
	totalTime = totalTime + time_taken
	if flagLegitimate == 0:
		sheet.cell(row = i, column = 2).value = "Phishing"
		phishCount = phishCount + 1

	if flagLegitimate == 1:
		sheet.cell(row = i, column = 2).value = "Legitimate"
		legitCount = legitCount + 1

	sheet.cell(row = i, column = 3).value = time_taken 
	sheet.cell(row = i, column = 4).value = matchFreq
	sheet.cell(row = i, column = 5).value =	totalLinks
	sheet.cell(row = i, column = 6).value = rankTld
	i = i + 1
	time.sleep(2)


sheet.cell(row = i, column = 1).value = str(phishCount + legitCount)
sheet.cell(row = i, column = 2).value = str(phishCount)
sheet.cell(row = i, column = 3).value = str(legitCount)
try:
	sheet.cell(row = i, column = 4).value = str(totalTime / (phishCount + legitCount))
except:
	errorCount = errorCount + 1


book.save('test1.xlsx')

#353035069270211
