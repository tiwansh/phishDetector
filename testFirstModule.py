import openpyxl
import time
import titleSearch
import returnFirstOther
'''
limit = raw_input("Enter limit : ")
limit = int(limit)
print type(limit)
'''

#Filename.xlsx
book = openpyxl.load_workbook('test1.xlsx')

sheet = book.active
i = 2
s2 = sheet.cell(row = i, column = 1)
'''print s2.value

i = i + 1
s2 = sheet.cell(row = i, column = 1)
print s2.value
'''

phishCount = 0
legitCount = 0
totalTime = 0
errorCount = 0

while 1:
	#print "Here"
	s2 = sheet.cell(row = i, column = 1)
	if s2.value is None:
		break
	url = str(s2.value)


	#url = "http://www." + url
	print (url)
	#i = i + 1
	
	'''title = titleSearch.returnTitle(url)
	if(title is not ""):
		print "Title is : " + str(title)
		#flag_titleFound = True
	else:
		print ("Title not found!")
	'''

	start_time = time.time()
	#verdict = srcIPMatch.ipMatches(url, 3)
	try:
		#print "gone"
		verdict =  returnFirstOther.isLegit(url)
	except:
		i = i + 1
		continue
	time_taken = time.time() - start_time
	totalTime = totalTime + time_taken
	if verdict == 0:
		sheet.cell(row = i, column = 2).value = "Phishing"
		phishCount = phishCount + 1

	if verdict == 1:
		sheet.cell(row = i, column = 2).value = "Legitimate"
		legitCount = legitCount + 1

	sheet.cell(row = i, column = 3).value = time_taken 
	i = i + 1
	time.sleep(30)
	book.save('test1.xlsx')


sheet.cell(row = i, column = 1).value = str(phishCount + legitCount)
sheet.cell(row = i, column = 2).value = str(phishCount)
sheet.cell(row = i, column = 3).value = str(legitCount)
try:
	sheet.cell(row = i, column = 4).value = str(totalTime / (phishCount + legitCount))
except:
	errorCount = errorCount + 1


book.save('test1.xlsx')

#353035069270211
