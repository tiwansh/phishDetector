''' Uses both approaches(Title + Domain) using other crawling function '''

''' Uses returnFinalOther.py   '''
''' Autoamted script
	1. Create an excel file striclty following the pattern - 
		First Row -------    URLs		Status		ResponseTime        
	2. Now fill the columns under URLs with the URLs of the the websites being tested for.
	3. Save the excel file with any name "anyName.xlsx"
	4. In the line no. of the code, change the line to book = openpyxl.load_workbook("anyName.xlsx")
	   i.e. put the name of the excel file which you want to test within the parantheses.
	5. Save the code
	6. Open terminal  
	7. Go to the directory containing this file excel.py
	8. Enter "python excel.py" and press enter
	9. Output is given on the screen and open the excel file anyName.xlsx to see the database. 
'''

import openpyxl
import returnFinalOther
import time

'''
limit = raw_input("Enter limit : ")
limit = int(limit)
print type(limit)
'''

#Filename.xlsx
book = openpyxl.load_workbook('test1.xlsx')

sheet = book.active
i = 2
s2 = sheet.cell(row = i, column = 1)
'''print s2.value

i = i + 1
s2 = sheet.cell(row = i, column = 1)
print s2.value
'''

phishCount = 0
blogCount = 0
legitCount = 0
totalTime = 0
errorCount = 0

while 1:
	s2 = sheet.cell(row = i, column = 1)
	if s2.value is None:
		break
	url = str(s2.value)


	#url = "http://www." + url
	#print (url)
	#i = i + 1
	
	start_time = time.time()
	try:
		verdict = returnFinalOther.isLegit(url)
	except:
		sheet.cell(row = i, column = 2).value = "Error"
		i = i + 1
		continue
	if verdict == -1:
		sheet.cell(row = i, column = 2).value = "Error"
		i = i + 1
		continue
	time_taken = time.time() - start_time
	totalTime = totalTime + time_taken
	if verdict == 0:
		sheet.cell(row = i, column = 2).value = "Phishing"
		phishCount = phishCount + 1

	if verdict == 1:
		sheet.cell(row = i, column = 2).value = "Legitimate"
		legitCount = legitCount + 1

	if verdict == 2:
		sheet.cell(row = i, column = 2).value = "Blogging"
		blogCount = blogCount + 1

	sheet.cell(row = i, column = 3).value = time_taken 
	i = i + 1
	time.sleep(30)
#print s2.valueS

sheet.cell(row = i, column = 1).value = str(blogCount + phishCount + legitCount)
sheet.cell(row = i, column = 2).value = str(phishCount)
sheet.cell(row = i, column = 3).value = str(legitCount)
sheet.cell(row = i, column = 4).value = str(blogCount)
sheet.cell(row = i, column = 5).value = str(totalTime / (blogCount + phishCount + legitCount))


book.save('test1.xlsx')

print "Total Sites Tested : " + str(blogCount + phishCount + legitCount)
print "Phishing Sites : " + str(phishCount)
print "Legitimate Sites : " + str(legitCount)
print "Blogging Sites : " + str(blogCount)
print "Average time taken : " + str(totalTime / (blogCount + phishCount + legitCount))




'''Append ka

sheet.cell(row=2, column=2).value = 2


last me book.save 
book.save('appending.xlsx')


'''