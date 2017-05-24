import openpyxl


book = openpyxl.load_workbook('lol.xlsx')

sheet = book.active

i = 2
s2 = sheet.cell(row = i, column = 1)

while (s2.value) is not None:
	s2 = sheet.cell(row = i, column = 1)
	print (s2.value)
	i = i + 1
	